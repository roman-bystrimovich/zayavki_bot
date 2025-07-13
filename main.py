import os
import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters, ConversationHandler
)
from dotenv import load_dotenv
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
import shutil
from datetime import datetime, date, timedelta
import calendar

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)


# Загрузка переменных окружения
load_dotenv()

# Переменные окружения для бота и почты
BOT_TOKEN = os.getenv("BOT_TOKEN")
EMAIL_LOGIN = os.getenv("EMAIL_LOGIN")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT"))
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")
TEMPLATE_PATH = "template.xlsx" # Убедитесь, что template.xlsx существует в той же директории

# Состояния для ConversationHandler
# Обновлено количество состояний до 19
PROJECT, OBJECT, NAME, UNIT, QUANTITY, MODULE, POSITION_DELIVERY_DATE, \
ATTACHMENT_CHOICE, FILE_INPUT, LINK_INPUT, \
CONFIRM_ADD_MORE, \
EDIT_MENU, SELECT_POSITION, EDIT_FIELD_SELECTION, EDIT_FIELD_INPUT, \
FINAL_CONFIRMATION, GLOBAL_DELIVERY_DATE_SELECTION, \
EDITING_UNIT, EDITING_MODULE = range(19)

# Глобальные переменные для хранения данных пользователя и предварительно определенных списков
user_state = {}
projects = ["Stadler", "Мотели"]
objects = ["Мерке", "Аральск", "Атырау", "Каркаролинск", "Семипалатинск"]
modules = [f"{i+1}" for i in range(18)]
units = ["м2", "м3", "шт", "компл", "л", "кг", "тн"]

def fill_excel(project, object_name, positions, user_full_name, telegram_id_or_username):
    """
    Заполняет Excel-файл данными, включая дату поставки для каждой позиции, проект, объект,
    а также информацию о пользователе, от которого пришла заявка.
    """
    today = datetime.today().strftime("%d.%m.%Y")
    # Изменено: Добавлено user_full_name в имя файла, заменены пробелы на подчеркивания
    sanitized_user_name = user_full_name.replace(" ", "_")
    filename = f"Заявка_{project}_{object_name}_{sanitized_user_name}_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    output_dir = "out"

    template_full_path = os.path.abspath(TEMPLATE_PATH)

    os.makedirs(output_dir, exist_ok=True)
    new_path = os.path.join(output_dir, filename)

    shutil.copy(template_full_path, new_path)
    wb = load_workbook(new_path)
    ws = wb.active

    ws['G2'] = today
    ws['G3'] = project
    ws['G4'] = object_name
    ws['G5'] = user_full_name
    ws['G6'] = telegram_id_or_username

    logger.info(f"Writing to Excel: G2={today}, G3={project}, G4={object_name}, G5={user_full_name}, G6={telegram_id_or_username}")


    row_start_data = 9
    for i, pos in enumerate(positions):
        row = row_start_data + i

        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=2).value = pos["name"]
        ws.cell(row=row, column=3).value = pos["unit"]
        ws.cell(row=row, column=4).value = pos["quantity"]
        ws.cell(row=row, column=5).value = pos.get("delivery_date", "Не указано") # Дата из позиции
        ws.cell(row=row, column=6).value = pos["module"]
        ws.cell(row=row, column=7).value = pos.get("link", "") # Добавлено поле для ссылки в Excel
        logger.info(f"Writing position {i+1} to Excel: {pos}")


    wb.save(new_path)
    logger.info(f"Excel file saved to: {new_path}")
    return new_path

async def send_email(chat_id, project, object_name, positions, user_full_name, telegram_id_or_username, context=None):
    """
    Отправляет сгенерированный Excel-файл по электронной почте,
    с возможностью прикрепления дополнительных файлов и ссылок, привязанных к позициям,
    а также информацией о пользователе.
    """
    msg = EmailMessage()
    msg["Subject"] = f"Заявка на снабжение: {project} - {object_name}"
    msg["From"] = EMAIL_LOGIN
    msg["To"] = EMAIL_RECEIVER

    email_body = "Во вложении заявка на снабжение.\n\n"
    email_body += f"Проект: {project}\n"
    email_body += f"Объект: {object_name}\n"
    email_body += f"От кого: {user_full_name}\n"
    email_body += f"Telegram ID: {telegram_id_or_username}\n\n"
    email_body += "Позиции:\n"
    
    # Списки для файлов и ссылок, которые будут прикреплены к письму
    files_to_attach = []
    links_in_email = []

    for i, p in enumerate(positions):
        pos_info = (
            f"{i+1}. Модуль: {p.get('module', 'N/A')} | Наименование: {p.get('name', 'N/A')} | "
            f"Ед.изм.: {p.get('unit', 'N/A')} | Количество: {p.get('quantity', 'N/A')} | "
            f"Дата поставки: {p.get('delivery_date', 'N/A')}"
        )
        if p.get('link'):
            pos_info += f" | Ссылка: {p['link']}"
            links_in_email.append(f"Позиция {i+1} ({p.get('name', 'N/A')}): {p['link']}")
        if p.get('file_data'):
            pos_info += f" | Файл: {p['file_data'].get('file_name', 'N/A')}"
            files_to_attach.append((i+1, p['file_data'])) # Сохраняем индекс позиции для логов
        email_body += pos_info + "\n"

    if links_in_email:
        email_body += "\nОтдельные ссылки для позиций:\n" + "\n".join(links_in_email) + "\n"

    msg.set_content(email_body)
    logger.info(f"Email body generated for chat_id {chat_id}: \n{email_body}")

    try:
        file_path = fill_excel(project, object_name, positions, user_full_name, telegram_id_or_username)
        with open(file_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(file_path),
            )
        logger.info(f"Excel file '{file_path}' прикреплен к письму.")
    except Exception as e:
        logger.error(f"Ошибка при создании или прикреплении Excel файла: {e}")

    # Прикрепление файлов, связанных с позициями
    if context:
        for pos_index, file_data in files_to_attach:
            try:
                file_id = file_data['file_id']
                file_name = file_data['file_name']
                mime_type = file_data['mime_type']

                telegram_file = await context.bot.get_file(file_id)
                file_bytes = await telegram_file.download_as_bytearray()

                msg.add_attachment(
                    file_bytes,
                    maintype=mime_type.split('/')[0],
                    subtype=mime_type.split('/')[1],
                    filename=f"Позиция_{pos_index}_{file_name}", # Уникальное имя файла
                )
                logger.info(f"Дополнительный файл '{file_name}' для позиции {pos_index} прикреплен к письму.")
            except Exception as e:
                logger.error(f"Ошибка при скачивании или прикреплении файла '{file_name}' для позиции {pos_index}: {e}")
                msg.set_content(msg.get_content() + f"\n\nВнимание: Не удалось прикрепить файл '{file_name}' для позиции {pos_index} из-за ошибки: {e}")


    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_LOGIN, EMAIL_PASSWORD)
            server.send_message(msg)
        logger.info(f"Письмо успешно отправлено на {EMAIL_RECEIVER}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при отправке письма: {e}")
        raise

# === Telegram Handlers ===

async def initial_message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает первое сообщение от пользователя (или когда диалог неактивен)
    и предлагает кнопку "Создать заявку".
    """
    keyboard = [[KeyboardButton("Создать заявку")]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=False, resize_keyboard=True)
    await update.message.reply_text(
        "Привет! Я бот для создания заявок. Нажмите кнопку, чтобы начать.",
        reply_markup=reply_markup
    )

async def start_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Начинает новый разговор (после нажатия кнопки "Создать заявку")
    и предлагает пользователю выбрать проект.
    """
    chat_id = update.effective_chat.id
    user = update.effective_user

    first_name = user.first_name if user.first_name else ""
    last_name = user.last_name if user.last_name else ""
    user_full_name = f"{first_name} {last_name}".strip()
    telegram_id_or_username = user.username if user.username else str(user.id)

    user_state[chat_id] = {
        "user_full_name": user_full_name,
        "telegram_id_or_username": telegram_id_or_username,
        "project": None,
        "object": None,
        "positions": [], # Позиции теперь могут содержать 'link' и 'file_data'
    }
    logger.info(f"User {user_full_name} ({telegram_id_or_username}) started conversation.")

    await update.message.reply_text("Начинаем создание заявки...", reply_markup=ReplyKeyboardRemove())

    keyboard = [[InlineKeyboardButton(p, callback_data=p)] for p in projects]
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите проект:", reply_markup=reply_markup)
    return PROJECT

async def project_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор проекта и предлагает выбрать объект."""
    query = update.callback_query
    await query.answer()

    user_state[query.message.chat.id]["project"] = query.data
    logger.info(f"Chat {query.message.chat.id}: Project selected - {query.data}")

    keyboard = [[InlineKeyboardButton(o, callback_data=o)] for o in objects]
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text("Выберите объект:", reply_markup=reply_markup)
    return OBJECT

async def object_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор объекта и запрашивает наименование позиции."""
    query = update.callback_query
    await query.answer()

    user_state[query.message.chat.id]["object"] = query.data
    logger.info(f"Chat {query.message.chat.id}: Object selected - {query.data}")
    await query.edit_message_text("Введите наименование позиции:")
    return NAME

async def name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает наименование позиции и предлагает выбрать единицу измерения."""
    user_state[update.effective_chat.id]["current"] = {"name": update.message.text}
    logger.info(f"Chat {update.effective_chat.id}: Position name entered - {update.message.text}")

    keyboard = [[InlineKeyboardButton(u, callback_data=u)] for u in units]
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите единицу измерения:", reply_markup=reply_markup)
    return UNIT

async def unit_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает выбор единицы измерения и запрашивает количество."""
    query = update.callback_query
    await query.answer()

    user_state[query.message.chat.id]["current"]["unit"] = query.data
    logger.info(f"Chat {query.message.chat.id}: Unit selected - {query.data}")
    await query.edit_message_text("Введите количество:")
    return QUANTITY

async def quantity_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает количество и предлагает выбрать модуль. Добавлена базовая валидация."""
    chat_id = update.effective_chat.id
    try:
        quantity = float(update.message.text)
        user_state[chat_id]["current"]["quantity"] = quantity
        logger.info(f"Chat {chat_id}: Quantity entered - {quantity}")
    except ValueError:
        logger.warning(f"Chat {chat_id}: Invalid quantity format - '{update.message.text}'")
        await update.message.reply_text("Неверный формат количества. Пожалуйста, введите число (например, 5 или 3.5):")
        return QUANTITY

    buttons_per_row = 5
    keyboard = []
    current_row = []
    for i, m in enumerate(modules):
        current_row.append(InlineKeyboardButton(m, callback_data=m))
        if (i + 1) % buttons_per_row == 0 or (i + 1) == len(modules):
            keyboard.append(current_row)
            current_row = []
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("К какому модулю относится позиция?", reply_markup=reply_markup)
    return MODULE

async def module_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор модуля и переходит к выбору даты поставки для текущей позиции.
    """
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat.id
    user_state[chat_id]["current"]["module"] = query.data # Сохраняем модуль
    logger.info(f"Chat {chat_id}: Module selected - {query.data}. Requesting delivery date for this position.")

    # Переходим к выбору даты поставки для текущей позиции
    current_date = date.today()
    # Используем префикс "POS_CAL_" для колбэков календаря позиций
    reply_markup = create_calendar_keyboard(current_date.year, current_date.month, prefix="POS_CAL_")
    await query.edit_message_text("Выберите желаемую дату поставки для этой позиции:", reply_markup=reply_markup)
    return POSITION_DELIVERY_DATE

async def process_position_calendar_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает нажатия на кнопки календаря для выбора даты поставки отдельной позиции.
    """
    query = update.callback_query
    await query.answer()

    data = query.data
    chat_id = query.message.chat.id

    if data.startswith("POS_CAL_NAV_"): # Обработка навигации
        parts = data.split('_')
        year = int(parts[3])
        month = int(parts[4])

        # Корректировка месяца и года
        if month > 12:
            month = 1
            year += 1
        elif month < 1:
            month = 12
            year -= 1

        reply_markup = create_calendar_keyboard(year, month, prefix="POS_CAL_")
        await query.edit_message_text("Выберите желаемую дату поставки для этой позиции:", reply_markup=reply_markup)
        return POSITION_DELIVERY_DATE

    elif data.startswith("POS_CAL_DATE_"): # Обработка выбора даты
        selected_date_str = data.replace("POS_CAL_DATE_", "")

        # Сохраняем дату в текущей позиции, но еще не добавляем в список позиций
        user_state[chat_id]["current"]["delivery_date"] = selected_date_str
        logger.info(f"Chat {chat_id}: Position delivery date selected - {selected_date_str}. Now asking about attachments.")

        # Предлагаем варианты прикрепления
        keyboard = [
            [InlineKeyboardButton("Прикрепить файл", callback_data="attach_file")],
            [InlineKeyboardButton("Прикрепить ссылку", callback_data="attach_link")],
            [InlineKeyboardButton("Продолжить без вложений", callback_data="no_attachment")]
        ]
        keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Теперь вы можете прикрепить файл или ссылку к этой позиции:", reply_markup=reply_markup)
        return ATTACHMENT_CHOICE # Переход в новое состояние

    elif data == "POS_CAL_CANCEL":
        logger.info(f"Chat {chat_id}: Position calendar date selection cancelled.")
        # Если пользователь отменяет выбор даты для позиции,
        # текущая неполная позиция должна быть удалена,
        # и пользователь возвращается в меню редактирования.
        if "current" in user_state[chat_id]:
            del user_state[chat_id]["current"]
        await query.edit_message_text("Выбор даты для позиции отменен. Вы можете добавить позицию снова или продолжить.")
        return await edit_menu_handler(update, context) # Вернуться в меню редактирования

    return POSITION_DELIVERY_DATE

async def attachment_choice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор пользователя по прикреплению файла, ссылки или продолжению без вложений.
    """
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    data = query.data

    if data == "attach_file":
        await query.edit_message_text("Пожалуйста, **отправьте мне файл** (как документ) для этой позиции.",
                                      reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
        return FILE_INPUT
    elif data == "attach_link":
        await query.edit_message_text("Пожалуйста, **введите ссылку** для этой позиции.",
                                      reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
        return LINK_INPUT
    elif data == "no_attachment":
        # Добавляем текущую позицию в список позиций, т.к. вложений не будет
        user_state[chat_id]["positions"].append(user_state[chat_id]["current"])
        logger.info(f"Chat {chat_id}: Position added without attachments: {user_state[chat_id]['current']}")
        del user_state[chat_id]["current"] # Очищаем current после добавления

        keyboard = [
            [InlineKeyboardButton("Да", callback_data="yes"), InlineKeyboardButton("Нет", callback_data="no")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("Позиция добавлена. Добавить ещё позицию?", reply_markup=reply_markup)
        return CONFIRM_ADD_MORE
    else:
        await query.edit_message_text("Неизвестный выбор.")
        return ATTACHMENT_CHOICE # Stay in state

async def handle_file_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает получение файла и сохраняет его данные в текущую позицию.
    """
    chat_id = update.effective_chat.id
    
    if update.message.document:
        document = update.message.document
        user_state[chat_id]["current"]["file_data"] = {
            'file_id': document.file_id,
            'file_name': document.file_name,
            'mime_type': document.mime_type
        }
        logger.info(f"Chat {chat_id}: File '{document.file_name}' attached to current position.")
        await update.message.reply_text(f"Файл '{document.file_name}' успешно прикреплен.")
    else:
        logger.warning(f"Chat {chat_id}: Expected document but received something else for file input.")
        await update.message.reply_text("Это не похоже на файл-документ. Пожалуйста, отправьте файл (документ).")
        return FILE_INPUT # Stay in state if not a document

    # После прикрепления файла, предлагаем прикрепить ссылку или продолжить
    keyboard = []
    current_pos_data = user_state[chat_id]["current"]
    
    # Если ссылка еще не прикреплена, предлагаем прикрепить ссылку
    if not current_pos_data.get('link'):
        keyboard.append([InlineKeyboardButton("Прикрепить ссылку", callback_data="attach_link")])
    
    keyboard.append([InlineKeyboardButton("Продолжить без вложений", callback_data="no_attachment")])
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text("Что дальше?", reply_markup=reply_markup)
    return ATTACHMENT_CHOICE

async def handle_link_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает получение ссылки и сохраняет ее в текущую позицию.
    """
    chat_id = update.effective_chat.id
    link = update.message.text.strip()

    if link.startswith("http://") or link.startswith("https://"):
        user_state[chat_id]["current"]["link"] = link
        logger.info(f"Chat {chat_id}: Link '{link}' attached to current position.")
        await update.message.reply_text(f"Ссылка '{link}' успешно прикреплена.")
    else:
        logger.warning(f"Chat {chat_id}: Invalid link format for link input - '{link}'")
        await update.message.reply_text("Пожалуйста, введите корректную ссылку, начинающуюся с http:// или https://.")
        return LINK_INPUT # Stay in state if invalid link

    # После прикрепления ссылки, предлагаем прикрепить файл или продолжить
    keyboard = []
    current_pos_data = user_state[chat_id]["current"]

    # Если файл еще не прикреплен, предлагаем прикрепить файл
    if not current_pos_data.get('file_data'):
        keyboard.append([InlineKeyboardButton("Прикрепить файл", callback_data="attach_file")])
    
    keyboard.append([InlineKeyboardButton("Продолжить без вложений", callback_data="no_attachment")])
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text("Что дальше?", reply_markup=reply_markup)
    return ATTACHMENT_CHOICE


# --- ОБРАБОТЧИКИ ДЛЯ РЕДАКТИРОВАНИЯ ---

def get_positions_summary(positions):
    """Формирует читаемую сводку позиций, включая прикрепленные ссылки и файлы."""
    if not positions:
        return "Позиции отсутствуют."
    summary_lines = []
    for i, p in enumerate(positions):
        line = (
            f"{i+1}. Модуль: {p.get('module', 'N/A')} | Наименование: {p.get('name', 'N/A')} | "
            f"Ед.изм.: {p.get('unit', 'N/A')} | Количество: {p.get('quantity', 'N/A')} | "
            f"Дата поставки: {p.get('delivery_date', 'N/A')}"
        )
        if p.get('link'):
            line += f" | Ссылка: {p['link']}"
        if p.get('file_data'):
            line += f" | Файл: {p['file_data'].get('file_name', 'N/A')}"
        summary_lines.append(line)
    return "\n".join(summary_lines)

async def edit_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Отображает сводку текущих позиций и предлагает опции редактирования/удаления/продолжения.
    """
    chat_id = update.effective_chat.id
    state = user_state.get(chat_id, {})
    positions = state.get("positions", [])

    summary_text = f"Текущие позиции в заявке:\n{get_positions_summary(positions)}\n\n"

    keyboard = []
    if positions:
        keyboard.append([InlineKeyboardButton("Редактировать позицию", callback_data="edit_pos")])
        keyboard.append([InlineKeyboardButton("Удалить позицию", callback_data="delete_pos")])

    keyboard.append([InlineKeyboardButton("Продолжить", callback_data="continue_final_confirm")])
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])

    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(summary_text + "Выберите действие:", reply_markup=reply_markup)
    else:
        await context.bot.send_message(chat_id=chat_id, text=summary_text + "Выберите действие:", reply_markup=reply_markup)

    return EDIT_MENU

async def select_position_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Предлагает пользователю выбрать позицию по номеру для редактирования или удаления.
    """
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat.id
    action = query.data
    user_state[chat_id]['action_type'] = action

    positions = user_state[chat_id].get("positions", [])
    if not positions:
        await query.edit_message_text("В заявке нет позиций для редактирования или удаления. "
                                      "Нажмите 'Продолжить' или 'Отмена заявки'.",
                                      reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Продолжить", callback_data="continue_final_confirm")],
                                                                         [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
        return EDIT_MENU # Возвращаемся в EDIT_MENU

    keyboard = []
    buttons_per_row = 5
    current_row = []
    for i in range(len(positions)):
        button_text = f"{i+1}"
        current_row.append(InlineKeyboardButton(button_text, callback_data=f"select_pos_{i}"))
        if (i + 1) % buttons_per_row == 0 or (i + 1) == len(positions):
            keyboard.append(current_row)
            current_row = []

    keyboard.append([InlineKeyboardButton("Назад в меню", callback_data="back_to_edit_menu")])
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(f"Выберите номер позиции для { 'редактирования' if action == 'edit_pos' else 'удаления' }:\n"
                                  f"{get_positions_summary(positions)}", reply_markup=reply_markup)

    return SELECT_POSITION

async def process_selected_position(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбранную позицию (для редактирования или удаления).
    """
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat.id

    if query.data == "back_to_edit_menu":
        return await edit_menu_handler(update, context)

    try:
        selected_index = int(query.data.split('_')[2])
    except (IndexError, ValueError):
        await query.edit_message_text("Неверный выбор позиции. Пожалуйста, попробуйте снова.",
                                      reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад в меню", callback_data="back_to_edit_menu")],
                                                                         [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
        return SELECT_POSITION # Остаемся в SELECT_POSITION

    positions = user_state[chat_id].get("positions", [])
    if selected_index < 0 or selected_index >= len(positions):
        await query.edit_message_text("Выбрана несуществующая позиция. Пожалуйста, выберите номер из списка.",
                                      reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад в меню", callback_data="back_to_edit_menu")],
                                                                         [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
        return SELECT_POSITION

    action_type = user_state[chat_id].get('action_type')

    if action_type == 'delete_pos':
        deleted_pos = positions.pop(selected_index)
        logger.info(f"Chat {chat_id}: Position deleted - {deleted_pos.get('name', '')}")
        await query.edit_message_text(f"Позиция '{deleted_pos.get('name', '')}' удалена.\n\n"
                                      f"Текущие позиции:\n{get_positions_summary(positions)}")
        return await edit_menu_handler(update, context)
    elif action_type == 'edit_pos':
        user_state[chat_id]['editing_position_index'] = selected_index
        logger.info(f"Chat {chat_id}: Editing position index - {selected_index}")
        return await edit_field_selection_handler(update, context)
    else:
        logger.warning(f"Chat {chat_id}: Unknown action type in process_selected_position - {action_type}")
        await query.edit_message_text("Неизвестное действие. Пожалуйста, попробуйте снова.",
                                      reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Назад в меню", callback_data="back_to_edit_menu")],
                                                                         [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
        return await edit_menu_handler(update, context)

async def edit_field_selection_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Предлагает пользователю выбрать поле для редактирования в выбранной позиции.
    Теперь включает опции для прикрепления/редактирования ссылки и файла.
    """
    query = update.callback_query
    if query: await query.answer()

    chat_id = update.effective_chat.id
    selected_index = user_state[chat_id]['editing_position_index']
    current_pos = user_state[chat_id]['positions'][selected_index]

    summary_pos = (
        f"Редактирование позиции №{selected_index+1}:\n"
        f"Модуль: {current_pos.get('module', 'N/A')}\n"
        f"Наименование: {current_pos.get('name', 'N/A')}\n"
        f"Ед.изм.: {current_pos.get('unit', 'N/A')}\n"
        f"Количество: {current_pos.get('quantity', 'N/A')}\n"
        f"Дата поставки: {current_pos.get('delivery_date', 'N/A')}\n"
    )
    if current_pos.get('link'):
        summary_pos += f"Ссылка: {current_pos['link']}\n"
    if current_pos.get('file_data'):
        summary_pos += f"Файл: {current_pos['file_data'].get('file_name', 'N/A')}\n"
    summary_pos += "\n"


    keyboard = [
        [InlineKeyboardButton("Наименование", callback_data="edit_field_name")],
        [InlineKeyboardButton("Ед. изм.", callback_data="edit_field_unit")],
        [InlineKeyboardButton("Количество", callback_data="edit_field_quantity")],
        [InlineKeyboardButton("Модуль", callback_data="edit_field_module")],
        [InlineKeyboardButton("Дата поставки", callback_data="edit_field_delivery_date")],
        # Новые кнопки для прикрепления/изменения ссылки и файла к позиции
        [InlineKeyboardButton("Прикрепить/Изменить файл", callback_data="edit_field_attach_file")],
        [InlineKeyboardButton("Прикрепить/Изменить ссылку", callback_data="edit_field_attach_link")],
        [InlineKeyboardButton("Назад в меню", callback_data="back_to_edit_menu")],
        [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if query:
        await query.edit_message_text(summary_pos + "Выберите поле для редактирования:", reply_markup=reply_markup)
    else:
        await context.bot.send_message(chat_id=chat_id, text=summary_pos + "Выберите поле для редактирования:", reply_markup=reply_markup)

    return EDIT_FIELD_SELECTION

async def edit_field_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Принимает новое значение для выбранного поля и обновляет позицию.
    Обрабатывает текстовый ввод, файлы и перенаправляет на выбор кнопками для ед.изм. и модуля, даты.
    """
    chat_id = update.effective_chat.id
    current_state_data = user_state[chat_id]

    if update.callback_query:
        query = update.callback_query
        await query.answer()
        editing_field = query.data.replace("edit_field_", "")
        current_state_data['editing_field'] = editing_field
        logger.info(f"Chat {chat_id}: Editing field set to {editing_field}")

        if editing_field == 'delivery_date':
            current_date = date.today()
            reply_markup = create_calendar_keyboard(current_date.year, current_date.month, prefix="EDIT_CAL_")
            await query.edit_message_text("Выберите новую дату поставки:", reply_markup=reply_markup)
            return GLOBAL_DELIVERY_DATE_SELECTION
        elif editing_field == 'unit':
            keyboard = [[InlineKeyboardButton(u, callback_data=f"edit_unit_{u}")] for u in units]
            keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text("Выберите новую единицу измерения:", reply_markup=reply_markup)
            return EDITING_UNIT # Новое состояние
        elif editing_field == 'module':
            buttons_per_row = 5
            keyboard = []
            current_row = []
            for i, m in enumerate(modules):
                current_row.append(InlineKeyboardButton(m, callback_data=f"edit_module_{m}"))
                if (i + 1) % buttons_per_row == 0 or (i + 1) == len(modules):
                    keyboard.append(current_row)
                    current_row = []
            keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")])
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text("Выберите новый модуль:", reply_markup=reply_markup)
            return EDITING_MODULE # Новое состояние
        elif editing_field == 'attach_file':
            await query.edit_message_text("Пожалуйста, **отправьте мне файл** (как документ) для этой позиции.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
            return EDIT_FIELD_INPUT # Ждем файл
        elif editing_field == 'attach_link':
            await query.edit_message_text("Пожалуйста, **введите ссылку** для этой позиции.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]]))
            return EDIT_FIELD_INPUT # Ждем ссылку
        else: # name or quantity
            field_name_ru = {
                'name': 'наименование',
                'quantity': 'количество',
            }.get(editing_field, 'значение')
            await query.edit_message_text(f"Введите новое {field_name_ru}:")
            return EDIT_FIELD_INPUT

    # Обработка ввода (текст или файл)
    editing_position_index = current_state_data['editing_position_index']
    editing_field = current_state_data['editing_field']
    current_position = current_state_data['positions'][editing_position_index]

    if editing_field == 'quantity':
        try:
            new_value = float(update.message.text)
            current_position[editing_field] = new_value
            logger.info(f"Chat {chat_id}: Position field '{editing_field}' updated to '{new_value}' for index {editing_position_index}")
            await update.message.reply_text(f"Поле '{editing_field}' обновлено.")
            return await edit_menu_handler(update, context)
        except ValueError:
            logger.warning(f"Chat {chat_id}: Invalid quantity format for edit - '{update.message.text}'")
            await update.message.reply_text("Неверный формат количества. Пожалуйста, введите число (например, 5 или 3.5):")
            return EDIT_FIELD_INPUT
    elif editing_field == 'name':
        new_value = update.message.text.strip()
        current_position[editing_field] = new_value
        logger.info(f"Chat {chat_id}: Position field '{editing_field}' updated to '{new_value}' for index {editing_position_index}")
        await update.message.reply_text(f"Поле '{editing_field}' обновлено.")
        return await edit_menu_handler(update, context)
    elif editing_field == 'attach_file':
        if update.message.document:
            document = update.message.document
            current_position['file_data'] = {
                'file_id': document.file_id,
                'file_name': document.file_name,
                'mime_type': document.mime_type
            }
            logger.info(f"Chat {chat_id}: File '{document.file_name}' attached to position {editing_position_index}.")
            await update.message.reply_text(f"Файл '{document.file_name}' успешно прикреплен к позиции.")
            return await edit_menu_handler(update, context)
        else:
            logger.warning(f"Chat {chat_id}: Expected document but received something else for attach_file.")
            await update.message.reply_text("Это не похоже на файл-документ. Пожалуйста, отправьте файл (документ).")
            return EDIT_FIELD_INPUT # Остаемся в этом состоянии
    elif editing_field == 'attach_link':
        link = update.message.text.strip()
        if link.startswith("http://") or link.startswith("https://"):
            current_position['link'] = link
            logger.info(f"Chat {chat_id}: Link '{link}' attached to position {editing_position_index}.")
            await update.message.reply_text(f"Ссылка '{link}' успешно прикреплена к позиции.")
            return await edit_menu_handler(update, context)
        else:
            logger.warning(f"Chat {chat_id}: Invalid link format for attach_link - '{link}'")
            await update.message.reply_text("Пожалуйста, введите корректную ссылку, начинающуюся с http:// или https://.")
            return EDIT_FIELD_INPUT # Остаемся в этом состоянии
    else:
        logger.warning(f"Chat {chat_id}: Unexpected field or input type in edit_field_input_handler: field={editing_field}, update.message={update.message}")
        await update.message.reply_text("Произошла неизвестная ошибка при редактировании. Пожалуйста, попробуйте снова.")
        return await edit_menu_handler(update, context) # Вернуться в меню редактирования

async def process_edited_unit_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор единицы измерения при редактировании.
    """
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat.id
    selected_unit = query.data.replace("edit_unit_", "")

    editing_position_index = user_state[chat_id]['editing_position_index']
    user_state[chat_id]['positions'][editing_position_index]['unit'] = selected_unit
    logger.info(f"Chat {chat_id}: Position unit updated to '{selected_unit}' for index {editing_position_index}")

    await query.edit_message_text(f"Единица измерения обновлена на '{selected_unit}'.")
    return await edit_menu_handler(update, context)

async def process_edited_module_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор модуля при редактировании.
    """
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat.id
    selected_module = query.data.replace("edit_module_", "")

    editing_position_index = user_state[chat_id]['editing_position_index']
    user_state[chat_id]['positions'][editing_position_index]['module'] = selected_module
    logger.info(f"Chat {chat_id}: Position module updated to '{selected_module}' for index {editing_position_index}")

    await query.edit_message_text(f"Модуль обновлен на '{selected_module}'.")
    return await edit_menu_handler(update, context)


async def confirm_add_more_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает подтверждение добавления новой позиции или переход в меню редактирования.
    """
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat.id

    if query.data == "yes":
        logger.info(f"Chat {chat_id}: User wants to add more positions.")
        await query.edit_message_text("Введите наименование позиции:")
        return NAME
    else:
        logger.info(f"Chat {chat_id}: User finished adding positions, proceeding to edit menu.")
        return await edit_menu_handler(update, context)

# --- ОБРАБОТЧИКИ ДЛЯ КАЛЕНДАРЯ ---

def create_calendar_keyboard(year, month, prefix="CAL_"):
    """
    Генерирует Inline-клавиатуру для выбора даты.
    `prefix` используется для создания уникальных callback_data,
    например, "CAL_" для глобального выбора даты и "POS_CAL_" для даты позиции.
    """
    keyboard = []
    # Ряд 1: Навигация по месяцам и годам
    keyboard.append([
        InlineKeyboardButton("<<", callback_data=f"{prefix}NAV_{year-1}_{month}"),
        InlineKeyboardButton("<", callback_data=f"{prefix}NAV_{year}_{month-1 if month > 1 else 12}"),
        InlineKeyboardButton(f"{calendar.month_name[month]} {year}", callback_data="ignore"),
        InlineKeyboardButton(">", callback_data=f"{prefix}NAV_{year}_{month+1 if month < 12 else 1}"),
        InlineKeyboardButton(">>", callback_data=f"{prefix}NAV_{year+1}_{month}")
    ])

    # Ряд 2: Заголовки дней недели
    weekdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.append([InlineKeyboardButton(day, callback_data="ignore") for day in weekdays])

    # Дни календаря
    cal = calendar.Calendar()
    for week in cal.monthdayscalendar(year, month):
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="ignore"))
            else:
                current_day = date(year, month, day)
                row.append(InlineKeyboardButton(str(day), callback_data=f"{prefix}DATE_{current_day.isoformat()}"))
        keyboard.append(row)

    keyboard.append([InlineKeyboardButton("Отмена выбора даты", callback_data=f"{prefix}CANCEL")])
    keyboard.append([InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]) # Добавлена общая кнопка отмены

    return InlineKeyboardMarkup(keyboard)


async def request_global_delivery_date_calendar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Запрашивает глобальную дату поставки (если все позиции имеют одну дату).
    В текущей логике это может быть использовано для редактирования даты позиции,
    но это требует дополнительной адаптации.
    """
    current_date = date.today()
    reply_markup = create_calendar_keyboard(current_date.year, current_date.month, prefix="CAL_")

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text("Выберите желаемую общую дату поставки (если применимо):", reply_markup=reply_markup)
    else:
        await update.message.reply_text("Выберите желаемую общую дату поставки (если применимо):", reply_markup=reply_markup)

    return GLOBAL_DELIVERY_DATE_SELECTION # Новое состояние

async def process_global_calendar_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает нажатия на кнопки календаря (навигация или выбор даты) для глобальной даты.
    Используется также для редактирования даты поставки отдельной позиции.
    """
    query = update.callback_query
    await query.answer()

    data = query.data
    chat_id = query.message.chat.id

    if data.startswith("CAL_NAV_") or data.startswith("EDIT_CAL_NAV_"):
        parts = data.split('_')
        year = int(parts[2]) if data.startswith("CAL_NAV_") else int(parts[3])
        month = int(parts[3]) if data.startswith("CAL_NAV_") else int(parts[4])
        prefix = "CAL_" if data.startswith("CAL_NAV_") else "EDIT_CAL_" # Определяем префикс

        if month > 12:
            month = 1
            year += 1
        elif month < 1:
            month = 12
            year -= 1

        reply_markup = create_calendar_keyboard(year, month, prefix=prefix)
        await query.edit_message_text("Выберите дату:", reply_markup=reply_markup)
        return GLOBAL_DELIVERY_DATE_SELECTION # Остаемся в этом состоянии

    elif data.startswith("CAL_DATE_") or data.startswith("EDIT_CAL_DATE_"):
        selected_date_str = data.replace("CAL_DATE_", "").replace("EDIT_CAL_DATE_", "")

        if user_state[chat_id].get('editing_field') == 'delivery_date':
            # Если это редактирование даты для конкретной позиции
            editing_position_index = user_state[chat_id]['editing_position_index']
            user_state[chat_id]['positions'][editing_position_index]['delivery_date'] = selected_date_str
            logger.info(f"Chat {chat_id}: Position {editing_position_index} delivery date updated to {selected_date_str}")
            await query.edit_message_text(f"Дата поставки обновлена на {selected_date_str}.")
            # После редактирования возвращаемся в меню редактирования позиций
            return await edit_menu_handler(update, context)
        else:
            # Эта ветка, возможно, не будет достигнута при текущей логике, т.к. глобальная дата не используется.
            pass

        # После выбора даты всегда возвращаемся в меню редактирования
        return await edit_menu_handler(update, context)

    elif data == "CAL_CANCEL" or data == "EDIT_CAL_CANCEL":
        logger.info(f"Chat {chat_id}: Calendar date selection cancelled.")
        await query.edit_message_text("Выбор даты отменен.")
        # После отмены выбора даты, возвращаемся в меню редактирования
        return await edit_menu_handler(update, context)

    return GLOBAL_DELIVERY_DATE_SELECTION # Остаемся в этом состоянии

# --- ФИНАЛЬНОЕ ПОДТВЕРЖДЕНИЕ И ОТПРАВКА ---

async def show_final_summary_and_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Формирует финальное резюме заявки и запрашивает подтверждение отправки.
    """
    chat_id = update.effective_chat.id
    state = user_state[chat_id]

    positions_summary = get_positions_summary(state["positions"])

    full_summary = (
        f"Проект: {state['project']}\n"
        f"Объект: {state['object']}\n"
        f"От кого: {state.get('user_full_name', 'Неизвестно')}\n"
        f"Telegram ID: {state.get('telegram_id_or_username', 'Неизвестно')}\n\n"
        f"Позиции:\n{positions_summary}\n\n"
    )

    full_summary += "Отправить заявку на почту? (Да/Нет)"

    keyboard = [
        [InlineKeyboardButton("Да", callback_data="final_yes"), InlineKeyboardButton("Нет", callback_data="final_no")],
        [InlineKeyboardButton("Отмена заявки", callback_data="cancel_dialog")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.edit_message_text(full_summary, reply_markup=reply_markup)
    else:
        await update.message.reply_text(full_summary, reply_markup=reply_markup)

    return FINAL_CONFIRMATION

async def final_confirm_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает финальное подтверждение и отправляет заявку или отменяет ее."""
    query = update.callback_query
    await query.answer()

    chat_id = query.message.chat.id
    state = user_state[chat_id]

    if query.data == "final_yes":
        try:
            email_sent = await send_email(
                chat_id,
                state["project"],
                state["object"],
                state["positions"],
                state.get("user_full_name", "Неизвестно"),
                state.get("telegram_id_or_username", "Неизвестно"),
                context=context # Передаем context для скачивания файлов позиций
            )
            if email_sent:
                await query.edit_message_text("Заявка успешно отправлена на почту!")
            else:
                await query.edit_message_text("Заявка отправлена, но возникли проблемы при отправке письма. Пожалуйста, проверьте логи.")

            keyboard = [[KeyboardButton("Создать заявку")]]
            reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=False, resize_keyboard=True)
            await context.bot.send_message(chat_id=chat_id, text="Для создания новой заявки:", reply_markup=reply_markup)

            if chat_id in user_state:
                del user_state[chat_id]
        except Exception as e:
            logger.error(f"Ошибка в final_confirm_handler: {e}")
            await query.edit_message_text(f"Произошла ошибка при отправке заявки: {e}\nПожалуйста, попробуйте еще раз позднее.")
            keyboard = [[KeyboardButton("Создать заявку")]]
            reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=False, resize_keyboard=True)
            await context.bot.send_message(chat_id=chat_id, text="Для создания новой заявки:", reply_markup=reply_markup)

        return ConversationHandler.END
    else: # query.data == "final_no"
        await query.edit_message_text("Отправка заявки отменена.")
        keyboard = [[KeyboardButton("Создать заявку")]]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=False, resize_keyboard=True)
        await context.bot.send_message(chat_id=chat_id, text="Для создания новой заявки:", reply_markup=reply_markup)
        if chat_id in user_state:
            del user_state[chat_id]
        return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отменяет текущий разговор и очищает состояние пользователя."""
    chat_id = update.effective_chat.id
    
    # Сначала отправляем сообщение об отмене и убираем любую ReplyKeyboard, если она есть
    await context.bot.send_message(chat_id=chat_id, text='Диалог отменён.', reply_markup=ReplyKeyboardRemove())

    # Если это CallbackQuery, отвечаем на него и пытаемся удалить inline-клавиатуру из предыдущего сообщения
    if update.callback_query:
        query = update.callback_query
        await query.answer() # Отвечаем на запрос, чтобы убрать "часы" на кнопке
        try:
            # Пытаемся удалить инлайн-клавиатуру, если она присутствует в сообщении
            if query.message and query.message.reply_markup and query.message.reply_markup.inline_keyboard:
                await query.edit_message_reply_markup(reply_markup=None)
        except Exception as e:
            logger.warning(f"Failed to edit message to remove inline keyboard after cancel: {e}")

    # После всего, отправляем новую кнопку "Создать заявку"
    keyboard = [[KeyboardButton("Создать заявку")]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=False, resize_keyboard=True)
    await context.bot.send_message(chat_id=chat_id, text="Для создания новой заявки:", reply_markup=reply_markup)

    # Очищаем состояние пользователя
    if chat_id in user_state:
        del user_state[chat_id]
        logger.info(f"Chat {chat_id} state cleared after cancel.")
        
    return ConversationHandler.END

async def unknown(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ответ на неизвестные команды или сообщения, не относящиеся к текущему диалогу."""
    chat_id = update.effective_chat.id
    keyboard = [[KeyboardButton("Создать заявку")]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=False, resize_keyboard=True)

    if update.message:
        await update.message.reply_text("Извините, я не понял вашу команду или сообщение. Пожалуйста, используйте кнопку 'Создать заявку' или начните заново командой /start.", reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.answer("Неизвестное действие.")
        await context.bot.send_message(chat_id=chat_id, text=".", reply_markup=reply_markup)


def main():
    """Основная функция для запуска бота."""
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & filters.Regex("^Создать заявку$"), start_conversation)],
        states={
            PROJECT: [
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
                CallbackQueryHandler(project_handler)
            ],
            OBJECT: [
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
                CallbackQueryHandler(object_handler)
            ],
            NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, name_handler),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            UNIT: [
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
                CallbackQueryHandler(unit_handler)
            ],
            QUANTITY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, quantity_handler),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            MODULE: [
                CallbackQueryHandler(module_handler, pattern="^(?:[1-9]|1[0-8])$"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            POSITION_DELIVERY_DATE: [
                CallbackQueryHandler(process_position_calendar_callback, pattern="^POS_CAL_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            # Новые состояния для прикрепления вложений
            ATTACHMENT_CHOICE: [
                CallbackQueryHandler(attachment_choice_handler, pattern="^(attach_file|attach_link|no_attachment)$"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            FILE_INPUT: [
                MessageHandler(filters.Document.ALL & ~filters.COMMAND, handle_file_input),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            LINK_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_link_input),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            CONFIRM_ADD_MORE: [
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
                CallbackQueryHandler(confirm_add_more_handler)
            ],

            EDIT_MENU: [
                CallbackQueryHandler(select_position_handler, pattern="^(edit_pos|delete_pos)$"),
                CallbackQueryHandler(show_final_summary_and_confirm, pattern="^continue_final_confirm$"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            SELECT_POSITION: [
                CallbackQueryHandler(process_selected_position, pattern="^select_pos_\\d+$"),
                CallbackQueryHandler(edit_menu_handler, pattern="^back_to_edit_menu$"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            EDIT_FIELD_SELECTION: [
                CallbackQueryHandler(edit_field_input_handler, pattern="^edit_field_"),
                CallbackQueryHandler(edit_menu_handler, pattern="^back_to_edit_menu$"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            EDIT_FIELD_INPUT: [
                MessageHandler(filters.TEXT | filters.Document.ALL & ~filters.COMMAND, edit_field_input_handler),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],

            EDITING_UNIT: [
                CallbackQueryHandler(process_edited_unit_selection, pattern="^edit_unit_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],
            EDITING_MODULE: [
                CallbackQueryHandler(process_edited_module_selection, pattern="^edit_module_"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],

            GLOBAL_DELIVERY_DATE_SELECTION: [
                CallbackQueryHandler(process_global_calendar_callback, pattern="^(CAL_|EDIT_CAL_)"),
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$")
            ],

            FINAL_CONFIRMATION: [
                CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
                CallbackQueryHandler(final_confirm_handler)
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CallbackQueryHandler(cancel, pattern="^cancel_dialog$"),
            MessageHandler(filters.COMMAND | filters.TEXT, unknown)
        ],
    )

    app.add_handler(conv_handler)

    app.add_handler(CommandHandler("start", initial_message_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, initial_message_handler))


    app.run_polling()

if __name__ == "__main__":
    main()